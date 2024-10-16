from openpyxl import load_workbook, Workbook

# Загрузить Excel файл
wb = load_workbook('download_file.xlsx')
ws = wb.active

# Создать новую рабочую книгу для отфильтрованных данных
filtered_wb = Workbook()
filtered_ws = filtered_wb.active

# Итерация по столбцам и фильтрация по условию
filtered_data = []

for col in ws.iter_cols(values_only=True):
    subject1 = 'МДК.07.01 Управление и автоматизация баз данных\nДавыдова Л.Б.'
    subject2 = 'МДК.11.01 Технология разработки и защиты баз данных\nДавыдова Л.Б.'
    if subject1 in col or subject2 in col:
        print(col)




# Сохранить отфильтрованные данные в новый Excel файл
filtered_wb.save('filtered_download_file.xlsx')

# Вывод отфильтрованных данных
print(filtered_data)
print(len(filtered_data))